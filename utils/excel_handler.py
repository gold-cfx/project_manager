#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
科研项目管理系统 - Excel处理工具模块
提供Excel文件的导入、导出和模板生成功能
"""
from typing import List, Dict, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

from utils.dict_utils import dict_utils

column_widths = [25, 15, 15, 15, 20, 25, 15, 18, 20, 15, 20, 15, 18, 18]


class ExcelTemplateGenerator:
    """Excel模板生成器"""

    PROJECT_HEADERS = [
        '项目名称',
        '项目负责人',
        '科室',
        '联系电话',
        '项目来源',
        '项目类型',
        '项目级别',
        '资助经费（万元）',
        '资助单位',
        '立项年度',
        '项目编号',
        '项目状态',
        '项目开始时间',
        '项目结束时间',
    ]

    @classmethod
    def generate_project_template(cls, file_path: str) -> bool:
        """生成项目录入模板"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "项目信息模板"

            # 设置标题样式
            title_font = Font(bold=True, size=12, color="FFFFFF")
            title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            title_alignment = Alignment(horizontal="center", vertical="center")

            # 添加标题行
            for col, header in enumerate(cls.PROJECT_HEADERS, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = title_font
                cell.fill = title_fill
                cell.alignment = title_alignment

            # 设置列宽

            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = width

            # 添加数据验证
            # 项目来源验证
            PROJECT_SOURCES = dict_utils.get_type_values(dict_utils.get_project_sources())
            source_validation = DataValidation(type="list", formula1=f'"{",".join(PROJECT_SOURCES)}"')
            source_validation.add(f'E2:E1000')
            ws.add_data_validation(source_validation)

            # 项目类型验证
            PROJECT_TYPES = dict_utils.get_type_values(dict_utils.get_project_types())
            type_validation = DataValidation(type="list", formula1=f'"{",".join(PROJECT_TYPES)}"')
            type_validation.add(f'F2:F1000')
            ws.add_data_validation(type_validation)

            # 项目级别验证
            PROJECT_LEVELS = dict_utils.get_type_values(dict_utils.get_project_levels())

            level_validation = DataValidation(
                type="list",
                formula1=f'"{",".join(PROJECT_LEVELS)}"'
            )
            level_validation.add(f'G2:G1000')
            ws.add_data_validation(level_validation)

            # 项目状态验证
            PROJECT_STATUS = dict_utils.get_type_values(dict_utils.get_project_status())
            status_validation = DataValidation(
                type="list",
                formula1=f'"{",".join(PROJECT_STATUS)}"'
            )
            status_validation.add(f'L2:L1000')
            ws.add_data_validation(status_validation)

            # 添加说明工作表
            instruction_ws = wb.create_sheet("填写说明")
            instructions = [
                ["填写说明", ""],
                ["", ""],
                ["项目名称", "必填，不超过255个字符"],
                ["项目负责人", "必填，不超过50个字符"],
                ["科室", "必填，不超过50个字符"],
                ["联系电话", "必填，格式：手机号或座机号"],
                ["项目来源", "从下拉列表中选择"],
                ["项目类型", "从下拉列表中选择"],
                ["项目级别", "从下拉列表中选择"],
                ["资助经费（万元）", "必填，数字格式，保留2位小数"],
                ["资助单位", "必填，不超过100个字符"],
                ["立项年度", "必填，格式：YYYY，如2024"],
                ["项目编号", "必填，不超过50个字符"],
                ["项目状态", "从下拉列表中选择"],
                ["项目开始时间", "必填，格式：YYYY-MM-DD"],
                ["项目结束时间", "必填，格式：YYYY-MM-DD，不能早于开始时间"],
            ]

            for row, (field, desc) in enumerate(instructions, 1):
                instruction_ws.cell(row=row, column=1, value=field)
                instruction_ws.cell(row=row, column=2, value=desc)

            # 设置说明工作表样式
            instruction_ws.column_dimensions['A'].width = 20
            instruction_ws.column_dimensions['B'].width = 50
            instruction_ws.cell(row=1, column=1).font = Font(bold=True, size=14)

            wb.save(file_path)
            return True

        except Exception as e:
            print(f"生成模板失败: {str(e)}")
            return False


class ExcelImporter:
    """Excel导入器"""

    @staticmethod
    def import_projects_from_excel(file_path: str) -> List[Dict[str, Any]]:
        """从Excel文件导入项目数据"""
        try:
            df = pd.read_excel(file_path, sheet_name="项目信息模板")

            # 检查必要的列
            required_columns = [
                '项目名称', '项目负责人', '科室', '联系电话',
                '项目来源', '项目类型', '项目级别', '资助经费（万元）',
                '资助单位', '立项年度', '项目编号', '项目状态',
                '项目开始时间', '项目结束时间'
            ]

            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                raise ValueError(f"缺少必要的列: {missing_columns}")

            # 处理空值
            df = df.dropna(subset=['项目名称'])  # 移除空行

            projects = []
            for _, row in df.iterrows():
                try:
                    project_data = {
                        'project_name': str(row['项目名称']).strip(),
                        'leader': str(row['项目负责人']).strip(),
                        'department': str(row['科室']).strip(),
                        'phone': str(row['联系电话']).strip(),
                        'project_source': str(row['项目来源']).strip(),
                        'project_type': str(row['项目类型']).strip(),
                        'level': str(row['项目级别']).strip(),
                        'funding_amount': float(row['资助经费（万元）']),
                        'funding_unit': str(row['资助单位']).strip(),
                        'approval_year': str(row['立项年度']).strip(),
                        'project_number': str(row['项目编号']).strip(),
                        'status': str(row['项目状态']).strip(),
                        'start_date': pd.to_datetime(row['项目开始时间']).date(),
                        'end_date': pd.to_datetime(row['项目结束时间']).date(),
                    }
                    projects.append(project_data)
                except Exception as e:
                    print(f"处理行数据失败: {str(e)}")
                    continue

            return projects

        except Exception as e:
            print(f"导入Excel失败: {str(e)}")
            raise


class ExcelExporter:
    """Excel导出器"""

    @staticmethod
    def export_projects_to_excel(projects: List[Dict[str, Any]], file_path: str) -> bool:
        """将项目数据导出到Excel"""
        try:
            if not projects:
                return False

            df = pd.DataFrame(projects)

            # 重命名列以匹配中文标题
            column_mapping = {
                'project_name': '项目名称',
                'leader': '项目负责人',
                'department': '科室',
                'phone': '联系电话',
                'project_source': '项目来源',
                'project_type': '项目类型',
                'level': '项目级别',
                'funding_amount': '资助经费（万元）',
                'funding_unit': '资助单位',
                'approval_year': '立项年度',
                'project_number': '项目编号',
                'status': '项目状态',
                'start_date': '项目开始时间',
                'end_date': '项目结束时间',
            }

            df = df.rename(columns=column_mapping)

            # 只保留需要的列
            df = df[[col for col in column_mapping.values() if col in df.columns]]

            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "项目信息"

            # 将DataFrame数据写入工作表
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # 设置样式
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # 应用样式到标题行
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 调整列宽
            for index, column in enumerate(ws.columns):
                max_length = column_widths[index]
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)
            return True

        except Exception as e:
            print(f"导出Excel失败: {str(e)}")
            return False
